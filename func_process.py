import os
import shutil
import json
from openpyxl import load_workbook
import exception as e
from exception import TotalProcessError as pr

def start_process(framework_path, project_createpath, project_foldername, 
        project_code, erp_initial, erp_startpath, erp_id, erp_pw):
    try:
        project_path = os.path.join(project_createpath,project_foldername) # 생성할 프로젝트 경로
        print("생성 프로젝트: ",project_path)
        if os.path.isdir(project_path):
            raise e.UsingCreateFolder(project_path)

        ### 1. 폴더 복사하기
        shutil.copytree(framework_path, project_path)


        ### 2. project.json 내부 변경하기
        jsonFile_path = os.path.join(project_path, "project.json")
        with open(jsonFile_path, "r", encoding='UTF8') as f:
            data = json.load(f)

        data["name"] = project_foldername

        with open(jsonFile_path, "w", encoding='UTF8') as f:
            json.dump(data, f)


        ### 3. config.xlsx 내부 변경하기
        configFile_path = os.path.join(project_path, "Data/Config.xlsx")

        # smaple masterlist파일명 추출
        wb = load_workbook(filename=configFile_path, data_only=True)
        arrSheets = wb.sheetnames
        wb_settings = wb[arrSheets[0]]          
        masterlist_filename = wb_settings["B16"].value
        wb.close


        wb = load_workbook(configFile_path)
        arrSheets = wb.sheetnames
        wb_settings = wb[arrSheets[0]]
            # Settings 시트
                # B2, B3 변경
        wb_settings["B2"] = project_foldername
        wb_settings["B3"] = project_code

            # Assets 시트
                # B2, B3, B4 변경
        wb_assets = wb[arrSheets[2]]
        wb_assets["B2"] = erp_initial
        wb_assets["B3"] = erp_initial+"_Credential"
        if erp_startpath != "":
            wb_assets["B4"] = erp_startpath

        wb.save(configFile_path)
        wb.close


        ## 4. MasterList 파일명 변경하기
        masterlist_path = os.path.join(project_path, "Data/Master/{0}".format(masterlist_filename))
        os.rename(masterlist_path, masterlist_path.replace(masterlist_filename, "{0}_MasterList.xlsx".format(project_code)))


        ### 5. Legacy 내부 변경하기
        sample = "[ERP(sample)]"
        legacy_sample_path = os.path.join(project_path, "Legacy/{0}".format(sample))
        os.rename(legacy_sample_path, legacy_sample_path.replace(sample, erp_initial))


        ### 6. Legacy/Common 내부 변경하기
        common_sample_path = os.path.join(project_path, "Legacy/Common/Comm_{0}".format(sample))
        os.rename(common_sample_path, common_sample_path.replace(sample, erp_initial))


        ### 7. Legacy/Common/ERP폴더 내부 변경하기
        common_erp_path = os.path.join(project_path, "Legacy/Common/Comm_{0}".format(erp_initial))
        for f in os.scandir(common_erp_path):
            with open(f, "r", encoding="UTF8") as f_f:
                f_content = f_f.read()

            f_content = f_content.replace(sample, erp_initial)

            with open(f, "w", encoding="UTF8") as f_f:
                f_f.write(f_content)

            os.rename(f.path, f.path.replace(sample, erp_initial))        


        ### 8. Test 내부 변경하기
        test_sample_path = os.path.join(project_path, "Test/{0}".format(sample))
        test_erp_path = test_sample_path.replace(sample, erp_initial)
        os.rename(test_sample_path, test_erp_path)


        ### 9. Test/ERP폴더 내부 변경하기
        for f in os.scandir(test_erp_path):
            with open(f, "r", encoding="UTF8") as f_f:
                f_content = f_f.read()

            f_content = f_content.replace(sample, erp_initial)

            with open(f, "w", encoding="UTF8") as f_f:
                f_f.write(f_content)

            os.rename(f.path, f.path.replace(sample, erp_initial))

        
        ### 10. 자격증명 등록하기
        if erp_initial != "" and erp_pw != "":
            os.system("cmdkey /generic:{0}_Credential /user:{1} /pass:{2}".format(erp_initial, erp_id, erp_pw))
    


    except FileExistsError as err:
        return (pr(err).title, pr(err).msg)
    except FileNotFoundError as err:
        return (pr(err).title, pr(err).msg)
    except e.NotExistPath as err:
        return (pr(err).title, pr(err).msg)
    except e.UsingCreateFolder as err:
        return (pr(err).title, pr(err).msg)
    except Exception as err:
        return (pr(err).title, pr(err).msg)


    return (True,"성공")

